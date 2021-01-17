#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup

import pymysql

from selenium import webdriver
from time import *

import openpyxl

driver = webdriver.Chrome('C:/Users/s/Desktop/20-1/DB/project/chromedriver.exe')
driver.get('http://www.nktech.net/inform/term/term_l.jsp')

wb = openpyxl.Workbook()
sheet = wb.active

ten = 0
for i in range(100):
    html = driver.page_source
    bsObject = BeautifulSoup(html, 'html.parser')
    table =bsObject.find('table', class_="tbd01 hover")
    trs = table.find_all('tr')
    for idx, tr in enumerate(trs):
        if idx>0:
            tbs = tr.find_all('td')
            sheet.cell(row=idx+1+ten, column=2).value = tbs[0].text
            sheet.cell(row=idx+1+ten, column=3).value = tbs[1].text
            sheet.cell(row=idx+1+ten, column=4).value = tbs[2].text
            sheet.cell(row=idx+1+ten, column=5).value = ' '.join(str(tbs[3].text).split())
    ten = ten + 10
    sleep(3)

wb.save('data.xlsx')
